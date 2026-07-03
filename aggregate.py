#!/usr/bin/env python3
"""フロンタル予算実績ダッシュボード — データ集計スクリプト"""

import openpyxl
import csv
import json
import math
import os
import sys
from datetime import datetime
from pathlib import Path

# ======================================================
# パス設定（役割分担: データ=ローカル / 成果物=Dropbox / 認証=環境変数）
#   環境変数で上書き可能。コードを編集せず配置先だけ変えられる。
#   - FRONTAL_DATA_DIR : 生データ置き場（既定=このスクリプト直下 data/。ローカル推奨）
#   - FRONTAL_OUTBOX   : 成果物出力先（既定=~/Dropbox/kuroda_work/outbox。Dropbox推奨）
#   ※認証(PAT等)はコード・Dropboxに置かず、環境変数/資格情報マネージャで扱うこと。
# ======================================================
_HERE = Path(__file__).parent
DATA_DIR = Path(os.environ.get("FRONTAL_DATA_DIR", _HERE / "data"))
OUTBOX_DIR = Path(os.environ.get("FRONTAL_OUTBOX", Path.home() / "Dropbox" / "kuroda_work" / "outbox"))

# ======================================================
# 列リゾルバ — DriveDoorのExcelヘッダー名から列位置を動的解決する。
#   旧実装は row[89]/row[100] 等の固定index依存で、DriveDoor側の
#   列構成が変わると沈黙して誤集計するリスクがあった。
#   ヘッダー名で解決し、見つからない場合のみ従来の既定indexへ
#   フォールバックしつつ WARN を出す（沈黙の誤りを防ぐ）。
#
#   注意: aliases はDriveDoorの実ヘッダー名に合わせて調整すること。
#   実出力のヘッダーと違う場合、WARNが出た項目のaliasを追記する。
# ======================================================

# logical field -> (ヘッダー名候補, 既定index[0始まり])
COLUMN_SPECS = {
    'date':    (['運行日', '運行年月日', '日付', '実車日'], 0),
    'haisha':  (['配車先区分', '配車区分', '自社傭車区分', '自社・傭車', '区分'], 7),
    'vehicle': (['車両', '車両名', '号車', '車番', '車両表示名', '自動車表示名'], 27),
    'yosha':   (['傭車先', '傭車先名', '傭車会社', '傭車先会社名', '傭車先名称'], 40),
    'sales':   (['請求金額', '売上金額', '請求額', '運賃'], 89),
    'cost':    (['支払金額', '傭車費', '支払額', '外注費', '傭車支払金額'], 100),
}


def _norm(s):
    """ヘッダー比較用の正規化（前後空白・全半角スペース除去）。"""
    if s is None:
        return ''
    return str(s).strip().replace(' ', '').replace('\u3000', '')


def resolve_columns(ws, specs=COLUMN_SPECS):
    """ワークシートのヘッダー行(1行目)から logical field -> 0始まり列index を解決。
    見つからない項目は既定indexへフォールバックし、WARNをstderrに出す。
    """
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_map = {}
    for idx, cell in enumerate(header):
        n = _norm(cell)
        if n and n not in header_map:
            header_map[n] = idx

    resolved = {}
    for field, (aliases, fallback) in specs.items():
        found = next((header_map[_norm(a)] for a in aliases if _norm(a) in header_map), None)
        if found is None:
            print(f"[WARN] 列『{field}』をヘッダーから特定できませんでした。"
                  f"既定位置 index={fallback} を使用します。"
                  f"（DriveDoorの実ヘッダー名を確認し COLUMN_SPECS['{field}'] のaliasに追加してください）",
                  file=sys.stderr)
            found = fallback
        resolved[field] = found
    return resolved


def _cell(row, idx):
    """行タプルから安全に値を取り出す（範囲外はNone）。"""
    return row[idx] if idx is not None and len(row) > idx else None


# シクロ（FJS利用貨物）判定。手順書STEP2準拠:
#   表記ゆれ「シクロ」「SICRO」「SICURO」の3パターンで照合する。
#   ※ "SICURO" は "SICRO" を含まないため、部分一致1本では取りこぼす。
_SICRO_PATTERNS = ('シクロ', 'SICRO', 'SICURO')


def is_sicro(yosha_name):
    if not yosha_name:
        return False
    s = str(yosha_name).upper()
    return any(p.upper() in s for p in _SICRO_PATTERNS)

# --- 日報Excel読み取り ---
def read_nippo(filepath, office_type):
    """
    日報Excelを読み取り、月別・区分別に売上・原価を集計する。
    office_type: 'honsha', 'kyoto', 'fjs'

    本社・京都: 配車先区分(Col H)が「自社」→ 一般(ippan)、「傭車」→ 利用(riyo)
    FJS: すべて利用(riyo)に分類
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    cols = resolve_columns(ws)  # ヘッダー名から列位置を解決

    # {(category, month): total}  category = 'ippan' or 'riyo'
    result = {'ippan_sales': {}, 'ippan_cost': {}, 'riyo_sales': {}, 'riyo_cost': {}}

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = _cell(row, cols['date'])  # 運行日
        if date_val is None:
            continue

        # 日付パース
        if isinstance(date_val, datetime):
            month = date_val.month
        elif isinstance(date_val, str):
            try:
                dt = datetime.strptime(date_val.strip(), "%Y/%m/%d")
                month = dt.month
            except ValueError:
                continue
        else:
            continue

        # 区分判定
        haisha_raw = _cell(row, cols['haisha'])
        haisha = str(haisha_raw).strip() if haisha_raw else ''
        if office_type == 'fjs':
            # FJS: 自社 → 一般(京都に加算)
            # 傭車で傭車先=シクロシュプリーム → 利用(京都に加算)
            # その他傭車 → 除外(フロンタルに関係ない)
            yosha_raw = _cell(row, cols['yosha'])
            yosha_name = str(yosha_raw).strip() if yosha_raw else ''
            if haisha == '自社':
                category = 'ippan'
            elif haisha == '傭車' and is_sicro(yosha_name):
                category = 'riyo'
            else:
                continue  # フロンタルに関係ない傭車は除外
        elif office_type == 'kyoto':
            # 京都: 自社 → 一般、傭車 → 保留(除外)
            if haisha == '自社':
                category = 'ippan'
            else:
                continue  # 京都の傭車は保留
        else:
            # 本社: 配車先区分で判定
            if haisha == '傭車':
                category = 'riyo'
            else:
                category = 'ippan'  # 自社 or その他 → 一般

        # 請求金額 / 支払金額（ヘッダー名解決済みの列位置から取得）
        sales_raw = _cell(row, cols['sales'])
        cost_raw = _cell(row, cols['cost'])

        sales = float(sales_raw) if sales_raw is not None else 0
        cost = float(cost_raw) if cost_raw is not None else 0

        result[f'{category}_sales'][month] = result[f'{category}_sales'].get(month, 0) + sales
        result[f'{category}_cost'][month] = result[f'{category}_cost'].get(month, 0) + cost

        # 品質ガード: 利用貨物なのに支払金額(外注費)が空/0 → 利用粗利を過大に見せる。
        # 手順書STEP0-2の思想（異常値検出）に沿い、月別に件数と請求額を記録する。
        if category == 'riyo' and sales > 0 and cost == 0:
            _blank = result.setdefault('_riyo_cost_blank', {})
            b = _blank.setdefault(month, {'count': 0, 'sales': 0})
            b['count'] += 1
            b['sales'] += sales

    # 品質ガードの警告出力（利用計上なのに外注費が空のレコード）
    blanks = result.get('_riyo_cost_blank', {})
    if blanks:
        total_cnt = sum(b['count'] for b in blanks.values())
        total_sales = sum(b['sales'] for b in blanks.values())
        months_str = ', '.join(f"{m}月{b['count']}件" for m, b in sorted(blanks.items()))
        print(f"[DATA-WARN] {office_type}: 利用貨物なのに支払金額(外注費)が空のレコード "
              f"計{total_cnt}件（請求計 {round(total_sales/1000):,}千円）: {months_str}。"
              f"外注費が過少→利用粗利が過大に出ます。DriveDoorの支払入力漏れを確認してください。",
              file=sys.stderr)

    wb.close()
    return result


# --- 車両経費CSV読み取り ---
def read_sharyo_keihi(filepath):
    """車両経費CSVを読み取り、月別・費目別に集計する。"""
    # CSV「車両整備経費区分」→ ダッシュボードカテゴリ のマッピング
    # 本社固定費, 京都固定費, リース料, 倉庫賃料, 看板, 税金・保険・印紙代等 は
    # 固定費（CF_FIXED）に含まれるため除外
    category_map = {
        '燃料': 'nenryo',
        '通行料': 'kotsu',
        '固定車両費': 'sharyo',
        'タイヤ': 'sharyo',
        'オイル': 'sharyo',
        '修繕費': 'sharyo',
        '保険料': 'hoken',
        '法定福利費': 'jinken',
        '諸経費': 'jinken',
    }

    # {category: {month: total}}
    monthly_costs = {cat: {} for cat in ['nenryo', 'kotsu', 'sharyo', 'jinken', 'hoken']}

    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 発生年月日から月を取得
            date_str = row.get('発生年月日', '').strip()
            if not date_str:
                continue
            try:
                dt = datetime.strptime(date_str, "%Y/%m/%d")
                month = dt.month
            except ValueError:
                continue

            # 車両整備経費区分
            himoku = row.get('車両整備経費区分', '').strip()
            cat_key = category_map.get(himoku)
            if cat_key is None:
                continue

            # 経費金額
            amount_str = row.get('経費金額', '0').strip().replace(',', '')
            try:
                amount = float(amount_str)
            except ValueError:
                amount = 0

            monthly_costs[cat_key][month] = monthly_costs[cat_key].get(month, 0) + amount

    return monthly_costs


def round_sen(value):
    """円を千円に変換（四捨五入）"""
    if value is None:
        return None
    return round(value / 1000)


def main():
    # --- 日報読み取り ---
    honsha = read_nippo(DATA_DIR / "nippo_honsha.xlsx", "honsha")
    kyoto = read_nippo(DATA_DIR / "nippo_kyoto.xlsx", "kyoto")
    fjs = read_nippo(DATA_DIR / "nippo_fjs.xlsx", "fjs")

    # --- 集約: 全事業所のippan/riyoを合算 ---
    # 一般(ippan) = 本社自社 + 京都自社 + FJS自社
    # 利用(riyo) = 本社傭車 + 京都傭車 + FJS傭車(得意先フロンタルのみ)
    all_months = set()
    for data in [honsha, kyoto, fjs]:
        for key in data:
            all_months.update(data[key].keys())
    max_month = max(all_months) if all_months else 3

    # 月別集計（1〜12月）
    ippan_sales_act = []
    riyo_sales_act = []
    riyo_cost_act = []

    for m in range(1, 13):
        if m <= max_month:
            is_val = (honsha['ippan_sales'].get(m, 0) + kyoto['ippan_sales'].get(m, 0)
                      + fjs['ippan_sales'].get(m, 0))
            ippan_sales_act.append(round_sen(is_val))

            rs_val = (honsha['riyo_sales'].get(m, 0) + kyoto['riyo_sales'].get(m, 0)
                      + fjs['riyo_sales'].get(m, 0))
            rc_val = (honsha['riyo_cost'].get(m, 0) + kyoto['riyo_cost'].get(m, 0)
                      + fjs['riyo_cost'].get(m, 0))
            riyo_sales_act.append(round_sen(rs_val))
            riyo_cost_act.append(round_sen(rc_val))
        else:
            ippan_sales_act.append(None)
            riyo_sales_act.append(None)
            riyo_cost_act.append(None)

    # 粗利計算 (一般原価は後で車両経費CSVから算出するため、仮にNone)
    riyo_gross_act = []
    for m in range(12):
        if riyo_sales_act[m] is not None and riyo_cost_act[m] is not None:
            riyo_gross_act.append(riyo_sales_act[m] - riyo_cost_act[m])
        else:
            riyo_gross_act.append(None)

    # --- 車両経費CSV読み取り ---
    sharyo_costs = read_sharyo_keihi(DATA_DIR / "sharyokeihi.csv")

    cost_nenryo_act = []
    cost_kotsu_act = []
    cost_sharyo_act = []
    cost_jinken_act = []
    cost_hoken_act = []

    for m in range(1, 13):
        if m <= max_month:
            cost_nenryo_act.append(round_sen(sharyo_costs['nenryo'].get(m, 0)))
            cost_kotsu_act.append(round_sen(sharyo_costs['kotsu'].get(m, 0)))
            cost_sharyo_act.append(round_sen(sharyo_costs['sharyo'].get(m, 0)))
            cost_jinken_act.append(round_sen(sharyo_costs['jinken'].get(m, 0)))
            cost_hoken_act.append(round_sen(sharyo_costs['hoken'].get(m, 0)))
        else:
            cost_nenryo_act.append(None)
            cost_kotsu_act.append(None)
            cost_sharyo_act.append(None)
            cost_jinken_act.append(None)
            cost_hoken_act.append(None)

    # 費目別原価がすべて0の月はnullにする（未入力扱い）
    for m in range(12):
        month_num = m + 1
        if month_num <= max_month:
            total_cost_items = sum(x or 0 for x in [
                cost_nenryo_act[m], cost_kotsu_act[m], cost_sharyo_act[m],
                cost_jinken_act[m], cost_hoken_act[m]
            ])
            if total_cost_items == 0:
                cost_nenryo_act[m] = None
                cost_kotsu_act[m] = None
                cost_sharyo_act[m] = None
                cost_jinken_act[m] = None
                cost_hoken_act[m] = None

    # 一般原価 = 車両経費CSV費目合計 (燃料+交通+車両+人件+保険)
    ippan_cost_act = []
    ippan_gross_act = []
    for m in range(12):
        cost_items = [cost_nenryo_act[m], cost_kotsu_act[m], cost_sharyo_act[m],
                      cost_jinken_act[m], cost_hoken_act[m]]
        if any(v is not None for v in cost_items):
            total_cost = sum(v or 0 for v in cost_items)
            ippan_cost_act.append(total_cost)
            if ippan_sales_act[m] is not None:
                ippan_gross_act.append(ippan_sales_act[m] - total_cost)
            else:
                ippan_gross_act.append(None)
        else:
            ippan_cost_act.append(None)
            ippan_gross_act.append(None)

    # --- 出力 ---
    result = {
        'ippan_sales_act': ippan_sales_act,
        'ippan_cost_act': ippan_cost_act,
        'ippan_gross_act': ippan_gross_act,
        'riyo_sales_act': riyo_sales_act,
        'riyo_cost_act': riyo_cost_act,
        'riyo_gross_act': riyo_gross_act,
        'cost_nenryo_act': cost_nenryo_act,
        'cost_kotsu_act': cost_kotsu_act,
        'cost_sharyo_act': cost_sharyo_act,
        'cost_jinken_act': cost_jinken_act,
        'cost_hoken_act': cost_hoken_act,
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))

    # サマリー表示
    print("\n--- サマリー ---")
    for m in range(max_month):
        mn = m + 1
        ig = ippan_gross_act[m]
        rg = riyo_gross_act[m]
        ig_str = f"{ig:,}" if ig is not None else "算出不可"
        rg_str = f"{rg:,}" if rg is not None else "算出不可"
        print(f"  {mn}月: 一般売上={ippan_sales_act[m]:,} 一般粗利={ig_str} 利用粗利={rg_str}")


if __name__ == "__main__":
    main()
