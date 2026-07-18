#!/usr/bin/env python3
"""フロンタル 予算実績 Excelレポート生成（3シート）— 復元版
旧 Downloads/_スキル管理/frontal_excel_report.py の後継。原本は旧PC廃棄で消失したため、
実成果物『フロンタル_予算実績_20260626.xlsx』を正解見本として、
検証済み集計エンジン（aggregate.py）から再構築した。

出力: フロンタル_予算実績_YYYYMMDD.xlsx
シート: 月次実績サマリー / 費目別原価 / 累計進捗
改良点（原本比）: 実績配列の手編集をやめ、data/ の生データから自動集計。
"""
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from aggregate import (read_nippo, read_sharyo_keihi, round_sen, DATA_DIR, OUTBOX_DIR)

BASE = Path(__file__).parent

# ===== 設定ノブ =====================================================
CF_FIXED = 16000                      # 見本準拠（固定費+リース+金融）。手順書の16,500ではない
LATEST_MONTH_IS_MTD = True            # 最新月はMTDとして確定から除外
# 2026-07-17 クライアント確定：シクロは4月〜「京都日報」ベースで計上（京都請求=売上/京都支払=外注費）。
# これにより4〜5月のFJS外注費の確定値継承は不要になった（京都側に実支払があるため）。継承は無効化。
INHERIT_CONFIRMED_GAICHU = False
CONFIRMED_GAICHU = {}   # 旧: {4:19003,5:20194}。新ルール(4月〜京都計上)で不要
SOKUHO_MONTHS = set()                 # 速報月（黄背景+⚡）。確定でも実測でもない月を入れる

# ===== 予算配列（千円・12ヶ月）見本から抽出 =========================
ippan_sales_bud = [40260, 41100, 52700, 55550, 52000, 58850, 62850, 56800, 62850, 62850, 62850, 56800]
ippan_cost_bud  = [31055, 32419, 40040, 42303, 39457, 45086, 48168, 43105, 48168, 48168, 48168, 43105]
ippan_gross_bud = [9205, 8681, 12660, 13247, 12543, 13764, 14682, 13695, 14682, 14682, 14682, 13695]
riyo_sales_bud  = [16100] * 12
riyo_cost_bud   = [15014] * 12
riyo_gross_bud  = [1086] * 12
nenryo_bud = [9710, 9830, 13110, 13890, 12060, 14470, 15290, 12840, 15290, 15290, 15290, 12840]
kotsu_bud  = [4405, 4405, 5015, 5295, 5545, 5595, 6015, 6015, 6015, 6015, 6015, 6015]
sharyo_bud = [2250, 2250, 2270, 2380, 2550, 2590, 2780, 2780, 2780, 2780, 2780, 2780]
hoken_bud  = [608, 618, 923, 983, 1020, 1017, 1080, 1100, 1080, 1080, 1080, 1100]
roumu_bud  = [14082, 15316, 18722, 19755, 18282, 21414, 23003, 20370, 23003, 23003, 23003, 20370]
cf_bud = [ig + rg - CF_FIXED for ig, rg in zip(ippan_gross_bud, riyo_gross_bud)]

# ===== 実績集計（検証済みエンジン） =================================
honsha = read_nippo(DATA_DIR / "nippo_honsha.xlsx", "honsha")
kyoto  = read_nippo(DATA_DIR / "nippo_kyoto.xlsx", "kyoto")
fjs    = read_nippo(DATA_DIR / "nippo_fjs.xlsx", "fjs")
sharyo = read_sharyo_keihi(DATA_DIR / "sharyokeihi.csv")

all_months = set()
for d in (honsha, kyoto, fjs):
    for k in d:
        all_months.update(d[k].keys() if isinstance(d[k], dict) else [])
max_month = max(all_months) if all_months else 3
CONFIRMED_MONTH = max_month - 1 if (LATEST_MONTH_IS_MTD and max_month > 1) else max_month

def msum(dicts, m):
    return sum(d.get(m, 0) for d in dicts)

def series(dicts):
    return [round_sen(msum(dicts, m)) if m <= max_month else None for m in range(1, 13)]

ippan_sales_act = series([honsha['ippan_sales'], kyoto['ippan_sales'], fjs['ippan_sales']])
riyo_sales_act  = series([honsha['riyo_sales'],  kyoto['riyo_sales'],  fjs['riyo_sales']])
riyo_cost_act   = series([honsha['riyo_cost'],   kyoto['riyo_cost'],   fjs['riyo_cost']])
nenryo_act = [round_sen(sharyo['nenryo'].get(m, 0)) if m <= max_month else None for m in range(1, 13)]
kotsu_act  = [round_sen(sharyo['kotsu'].get(m, 0))  if m <= max_month else None for m in range(1, 13)]
sharyo_act = [round_sen(sharyo['sharyo'].get(m, 0)) if m <= max_month else None for m in range(1, 13)]
hoken_act  = [round_sen(sharyo['hoken'].get(m, 0))  if m <= max_month else None for m in range(1, 13)]
roumu_act  = [round_sen(sharyo['jinken'].get(m, 0)) if m <= max_month else None for m in range(1, 13)]

# 外注費の確定値継承 / MTD月かつ継承値なしは None
if INHERIT_CONFIRMED_GAICHU:
    for m, v in CONFIRMED_GAICHU.items():
        if 1 <= m <= max_month:
            riyo_cost_act[m - 1] = v
BLANK = {}
for d in (honsha, kyoto, fjs):
    for m, b in d.get('_riyo_cost_blank', {}).items():
        a = BLANK.setdefault(m, {'count': 0, 'sales': 0})
        a['count'] += b['count']; a['sales'] += b['sales']
for m in list(BLANK):
    if m > CONFIRMED_MONTH and m not in CONFIRMED_GAICHU and 1 <= m <= max_month:
        riyo_cost_act[m - 1] = None

def add(a, b):
    return None if a is None or b is None else a + b
def sub(a, b):
    return None if a is None or b is None else a - b

ippan_cost_act = [None if any(x[i] is None for x in (nenryo_act, kotsu_act, sharyo_act, hoken_act, roumu_act))
                  else nenryo_act[i] + kotsu_act[i] + sharyo_act[i] + hoken_act[i] + roumu_act[i]
                  for i in range(12)]
ippan_gross_act = [sub(ippan_sales_act[i], ippan_cost_act[i]) for i in range(12)]
riyo_gross_act  = [sub(riyo_sales_act[i], riyo_cost_act[i]) for i in range(12)]
cf_act = [sub(add(ippan_gross_act[i], riyo_gross_act[i]), CF_FIXED) for i in range(12)]

# 確定月より後は実績を出さない（予算のみ表示）
def confirm_mask(arr):
    return [v if (i + 1) <= CONFIRMED_MONTH else None for i, v in enumerate(arr)]
for name in ['ippan_sales_act', 'ippan_cost_act', 'ippan_gross_act', 'riyo_sales_act',
             'riyo_cost_act', 'riyo_gross_act', 'cf_act',
             'nenryo_act', 'kotsu_act', 'sharyo_act', 'hoken_act', 'roumu_act']:
    globals()[name] = confirm_mask(globals()[name])

# ===== スタイル =====================================================
NAVY = "1E3A5F"; WHITE = "FFFFFF"
FILL = {'ippan': "2563EB", 'ippan2': "3B82F6", 'ippan_g': "047857",
        'riyo': "7C3AED", 'gaichu': "5B21B6", 'riyo_g': "059669", 'cf': "1E40AF"}
SLATE = "F1F5F9"; TOTAL = "EFF6FF"; SOKUHO = "FEF9C3"
BLUE_TXT = "0000FF"; GREEN = "10B981"; RED = "EF4444"; BLACK = "000000"; GREY = "6B7280"
NUM = '#,##0;▲#,##0;"-"'; PCT = '0.0%;▲0.0%;"-"'
CEN = Alignment(horizontal="center", vertical="center")
RGT = Alignment(horizontal="right")
thin = Side(style="thin", color="D1D5DB"); BORDER = Border(thin, thin, thin, thin)

def hdr(ws, coord, text, fill, merge=None):
    c = ws[coord]; c.value = text
    c.font = Font(bold=True, color=WHITE, size=10); c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = CEN; c.border = BORDER
    if merge:
        ws.merge_cells(merge)
        for row in ws[merge]:
            for cc in row:
                cc.fill = PatternFill("solid", fgColor=fill); cc.border = BORDER

def num(ws, coord, val, fill, kind='amt', color=None, bold=False):
    c = ws[coord]; c.value = val
    c.number_format = PCT if kind == 'pct' else NUM
    c.font = Font(color=color or BLACK, bold=bold, size=10)
    c.fill = PatternFill("solid", fgColor=fill); c.alignment = RGT; c.border = BORDER

def goodbad(val, base, higher_good=True, pct=False):
    """達成率/差異の良し悪し色。costは higher_good=False。"""
    if val is None:
        return BLACK
    good = (val >= (base if pct else 0)) == higher_good
    return GREEN if good else RED

def rate(act, bud):
    return act / bud if (act is not None and bud) else None

wb = Workbook()

# ============ シート1: 月次実績サマリー ============
ws = wb.active; ws.title = "月次実績サマリー"
ws["A1"] = "フロンタル 予算実績サマリー 2026年度（単位：千円）"
ws["A1"].font = Font(bold=True, color=WHITE, size=13); ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:V1")
for row in ws["A1:V1"]:
    for cc in row:
        cc.fill = PatternFill("solid", fgColor=NAVY)
note = f"データ取得日: {datetime.now().strftime('%Y/%m/%d')}\u3000確定=1〜{CONFIRMED_MONTH}月（{CONFIRMED_MONTH+1}月はMTDのため月次未反映）\u3000色: 青=一般 / 紫=利用 / 緑=粗利"
note += "\u3000シクロ利用計上: 1〜3月=FJS日報 / 4月〜=京都日報（京都請求=売上・京都支払=外注費）"
if BLANK:
    bm = ', '.join(f"{m}月{b['count']}件" for m, b in sorted(BLANK.items()))
    note += f"\u3000⚠️支払金額(外注費)が空: {bm}（利用粗利が過大の恐れ）"
ws["A2"] = note; ws["A2"].font = Font(color=GREY, size=9)

hdr(ws, "A3", "月", NAVY, "A3:A4")
groups = [("一般売上", 'ippan', 2, ['予算', '実績', '差異', '達成率']),
          ("一般原価", 'ippan2', 6, ['予算', '実績', '差異', '率']),
          ("一般粗利", 'ippan_g', 10, ['予算', '実績', '差異']),
          ("利用売上", 'riyo', 13, ['予算', '実績', '差異', '達成率']),
          ("外注費", 'gaichu', 17, ['予算', '実績']),
          ("利用粗利", 'riyo_g', 19, ['予算', '実績']),
          ("CF概算", 'cf', 21, ['予算', '実績'])]
from openpyxl.utils import get_column_letter as L
for title, fk, c0, subs in groups:
    hdr(ws, f"{L(c0)}3", title, FILL[fk], f"{L(c0)}3:{L(c0+len(subs)-1)}3")
    for j, s in enumerate(subs):
        hdr(ws, f"{L(c0+j)}4", s, FILL[fk])

MONTHS = [f"{m}月" for m in range(1, 13)]
for i, mlabel in enumerate(MONTHS):
    r = 5 + i; mnum = i + 1
    base_fill = SOKUHO if mnum in SOKUHO_MONTHS else (SLATE if mnum % 2 == 1 else WHITE)
    lbl = ("⚡" + mlabel) if mnum in SOKUHO_MONTHS else mlabel
    mc = ws.cell(r, 1, lbl); mc.alignment = CEN; mc.border = BORDER
    mc.fill = PatternFill("solid", fgColor=base_fill); mc.font = Font(size=10, color="374151")

    def block(c0, bud, act, higher_good=True, want_diff=True, want_rate=True, rate_label='達成率'):
        num(ws, f"{L(c0)}{r}", bud[i], base_fill)
        num(ws, f"{L(c0+1)}{r}", act[i], base_fill, color=BLUE_TXT)
        col = c0 + 2
        if want_diff:
            diff = sub(act[i], bud[i])
            num(ws, f"{L(col)}{r}", diff, base_fill, color=goodbad(diff, 0, higher_good))
            col += 1
        if want_rate:
            rt = rate(act[i], bud[i])
            num(ws, f"{L(col)}{r}", rt, base_fill, 'pct', goodbad(rt, 1.0, higher_good, pct=True))

    block(2, ippan_sales_bud, ippan_sales_act, True)
    block(6, ippan_cost_bud, ippan_cost_act, False)                 # 原価: 超過=悪
    block(10, ippan_gross_bud, ippan_gross_act, True, want_rate=False)
    block(13, riyo_sales_bud, riyo_sales_act, True)
    # 外注費/利用粗利/CF: 予算・実績のみ
    num(ws, f"Q{r}", riyo_cost_bud[i], base_fill)
    num(ws, f"R{r}", riyo_cost_act[i], base_fill, color=BLUE_TXT)
    num(ws, f"S{r}", riyo_gross_bud[i], base_fill)
    num(ws, f"T{r}", riyo_gross_act[i], base_fill, color=BLUE_TXT, bold=True)
    num(ws, f"U{r}", cf_bud[i], base_fill)
    num(ws, f"V{r}", cf_act[i], base_fill, color=BLUE_TXT)

ws.column_dimensions['A'].width = 8
for c in range(2, 23):
    ws.column_dimensions[L(c)].width = 9 if L(c) not in ('E', 'I', 'P') else 8

# ============ シート2: 費目別原価 ============
ws2 = wb.create_sheet("費目別原価")
ws2["A1"] = "一般原価（費目別）予算→実績 2026年度（単位：千円）"
ws2["A1"].font = Font(bold=True, color=WHITE, size=13); ws2["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws2.merge_cells("A1:S1")
for row in ws2["A1:S1"]:
    for cc in row:
        cc.fill = PatternFill("solid", fgColor=NAVY)
hdr(ws2, "A2", "月", NAVY, "A2:A3")
items = [("燃料費", nenryo_bud, nenryo_act), ("交通費", kotsu_bud, kotsu_act),
         ("車両費", sharyo_bud, sharyo_act), ("保険料", hoken_bud, hoken_act),
         ("労務費", roumu_bud, roumu_act)]
# 合計
total_bud = [ippan_cost_bud[i] for i in range(12)]
total_act = ippan_cost_act
cols2 = {}
for k, (name, _, _) in enumerate(items):
    c0 = 2 + k * 3
    cols2[name] = c0
    hdr(ws2, f"{L(c0)}2", name, FILL['ippan2'], f"{L(c0)}2:{L(c0+2)}2")
    for j, s in enumerate(['予算', '実績', '率']):
        hdr(ws2, f"{L(c0+j)}3", s, FILL['ippan2'])
hdr(ws2, "Q2", "合計", NAVY, "Q2:S2")
for j, s in enumerate(['予算', '実績', '率']):
    hdr(ws2, f"{L(17+j)}3", s, NAVY)

for i, mlabel in enumerate(MONTHS):
    r = 4 + i; mnum = i + 1
    base_fill = SLATE if mnum % 2 == 1 else WHITE
    mc = ws2.cell(r, 1, mlabel); mc.alignment = CEN; mc.border = BORDER
    mc.fill = PatternFill("solid", fgColor=base_fill); mc.font = Font(size=10, color="374151")
    for k, (name, bud, act) in enumerate(items):
        c0 = 2 + k * 3
        num(ws2, f"{L(c0)}{r}", bud[i], base_fill)
        num(ws2, f"{L(c0+1)}{r}", act[i], base_fill, color=BLUE_TXT)
        rt = rate(act[i], bud[i])
        num(ws2, f"{L(c0+2)}{r}", rt, base_fill, 'pct', goodbad(rt, 1.0, higher_good=False, pct=True))
    num(ws2, f"Q{r}", total_bud[i], base_fill, bold=True)
    num(ws2, f"R{r}", total_act[i], base_fill, color=BLUE_TXT, bold=True)
    rt = rate(total_act[i], total_bud[i])
    num(ws2, f"S{r}", rt, base_fill, 'pct', goodbad(rt, 1.0, higher_good=False, pct=True), bold=True)
ws2.column_dimensions['A'].width = 8
for c in range(2, 20):
    ws2.column_dimensions[L(c)].width = 9

# ============ シート3: 累計進捗 ============
ws3 = wb.create_sheet("累計進捗")
ws3["A1"] = "累計進捗サマリー 2026年度（単位：千円）"
ws3["A1"].font = Font(bold=True, color=WHITE, size=13); ws3["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws3.merge_cells("A1:J1")
for row in ws3["A1:J1"]:
    for cc in row:
        cc.fill = PatternFill("solid", fgColor=NAVY)
N = CONFIRMED_MONTH
hdr(ws3, "A2", "項目", NAVY, "A2:A3")
hdr(ws3, "B2", f"1〜{N-1}月（確定）", FILL['ippan'], "B2:D2")
hdr(ws3, "E2", f"{N}月（確定）", FILL['ippan'], "E2:G2")
hdr(ws3, "H2", f"1〜{N}月 累計", FILL['ippan'], "H2:J2")
for c0 in (2, 5, 8):
    for j, s in enumerate(['予算', '実績', '達成率']):
        hdr(ws3, f"{L(c0+j)}3", s, FILL['ippan'])

def cum(arr, lo, hi):
    vals = [arr[m - 1] for m in range(lo, hi + 1)]
    return None if any(v is None for v in vals) else sum(v for v in vals)

rows3 = [("一般売上", ippan_sales_bud, ippan_sales_act, True),
         ("一般原価", ippan_cost_bud, ippan_cost_act, False),
         ("一般粗利", ippan_gross_bud, ippan_gross_act, True),
         (None, None, None, None),
         ("利用売上", riyo_sales_bud, riyo_sales_act, True),
         ("外注費", riyo_cost_bud, riyo_cost_act, False),
         ("利用粗利", riyo_gross_bud, riyo_gross_act, True),
         (None, None, None, None),
         ("粗利合計",
          [ippan_gross_bud[i] + riyo_gross_bud[i] for i in range(12)],
          [add(ippan_gross_act[i], riyo_gross_act[i]) for i in range(12)], True)]
for k, (name, bud, act, hg) in enumerate(rows3):
    r = 4 + k
    if name is None:
        ws3.cell(r, 1, "\u3000")
        continue
    is_total = (name == "粗利合計")
    rfill = TOTAL if is_total else (SLATE if (k % 2 == 0) else WHITE)
    ac = ws3.cell(r, 1, name); ac.alignment = Alignment(horizontal="left")
    ac.font = Font(bold=is_total, color="374151", size=10); ac.fill = PatternFill("solid", fgColor=rfill); ac.border = BORDER
    for gi, (lo, hi) in enumerate([(1, N - 1), (N, N), (1, N)]):
        c0 = 2 + gi * 3
        b = cum(bud, lo, hi); a = cum(act, lo, hi); rt = rate(a, b)
        num(ws3, f"{L(c0)}{r}", b, rfill, bold=is_total)
        num(ws3, f"{L(c0+1)}{r}", a, rfill, color=BLUE_TXT, bold=is_total)
        num(ws3, f"{L(c0+2)}{r}", rt, rfill, 'pct', goodbad(rt, 1.0, hg, pct=True), bold=is_total)
ws3.column_dimensions['A'].width = 12
for c in range(2, 11):
    ws3.column_dimensions[L(c)].width = 14

# ===== 保存 =========================================================
stamp = datetime.now().strftime('%Y%m%d')
OUTBOX_DIR.mkdir(parents=True, exist_ok=True)
out = OUTBOX_DIR / f"フロンタル_予算実績_{stamp}.xlsx"  # 役割分担: 成果物はDropbox(既定)
wb.save(out)
print(f"Excel report saved: {out}")
if BLANK:
    bc = sum(b['count'] for b in BLANK.values())
    print(f"[PROVISIONAL] FJSシクロ等 外注費未入力 {bc}件 → 該当月は確定値継承で暫定表示")
