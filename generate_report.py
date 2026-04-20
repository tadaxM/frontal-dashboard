#!/usr/bin/env python3
"""フロンタル予算実績レポート — Excel出力スクリプト

aggregate.py の集計結果をExcelレポートとして出力する。
シート構成:
  Summary 累計KPI
  ① 一般売上・粗利
  ② 一般原価（費目別）
  ③ 利用売上・粗利
  ④ キャッシュフロー
  ⑤ 累計進捗
  ⑥ 拠点別粗利（本社・京都）
  ⑦ 車両別粗利（拠点ごと・降順）
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import sys
import importlib.util

# aggregate.py をモジュールとして読み込み
BASE = Path(__file__).parent
spec = importlib.util.spec_from_file_location("aggregate", BASE / "aggregate.py")
agg = importlib.util.module_from_spec(spec)

# aggregate.py の main() を書き換えて値を取得できるように import 実行
# aggregate.py は main() を print するだけなので、必要な関数を直接呼ぶ
sys.path.insert(0, str(BASE))
from aggregate import read_nippo, read_sharyo_keihi, round_sen, DATA_DIR

# === データ集計 ===
honsha = read_nippo(DATA_DIR / "nippo_honsha.xlsx", "honsha")
kyoto = read_nippo(DATA_DIR / "nippo_kyoto.xlsx", "kyoto")
fjs = read_nippo(DATA_DIR / "nippo_fjs.xlsx", "fjs")
sharyo = read_sharyo_keihi(DATA_DIR / "sharyokeihi.csv")

all_months = set()
for data in [honsha, kyoto, fjs]:
    for key in data:
        all_months.update(data[key].keys())
max_month = max(all_months) if all_months else 3

months = list(range(1, 13))


def monthly(sum_dict_list, month):
    return sum(d.get(month, 0) for d in sum_dict_list)


ippan_sales_act = []
ippan_cost_act = []
riyo_sales_act = []
riyo_cost_act = []
cost_nenryo_act = []
cost_kotsu_act = []
cost_sharyo_act = []
cost_jinken_act = []
cost_hoken_act = []

for m in months:
    if m <= max_month:
        ippan_sales_act.append(round_sen(
            monthly([honsha['ippan_sales'], kyoto['ippan_sales'], fjs['ippan_sales']], m)
        ))
        riyo_sales_act.append(round_sen(
            monthly([honsha['riyo_sales'], kyoto['riyo_sales'], fjs['riyo_sales']], m)
        ))
        riyo_cost_act.append(round_sen(
            monthly([honsha['riyo_cost'], kyoto['riyo_cost'], fjs['riyo_cost']], m)
        ))
        cost_nenryo_act.append(round_sen(sharyo['nenryo'].get(m, 0)))
        cost_kotsu_act.append(round_sen(sharyo['kotsu'].get(m, 0)))
        cost_sharyo_act.append(round_sen(sharyo['sharyo'].get(m, 0)))
        cost_jinken_act.append(round_sen(sharyo['jinken'].get(m, 0)))
        cost_hoken_act.append(round_sen(sharyo['hoken'].get(m, 0)))
    else:
        ippan_sales_act.append(None)
        riyo_sales_act.append(None)
        riyo_cost_act.append(None)
        cost_nenryo_act.append(None)
        cost_kotsu_act.append(None)
        cost_sharyo_act.append(None)
        cost_jinken_act.append(None)
        cost_hoken_act.append(None)

# 費目別原価がすべて0の月はnullに
for i, m in enumerate(months):
    if m <= max_month:
        total = sum(v or 0 for v in [cost_nenryo_act[i], cost_kotsu_act[i],
                                      cost_sharyo_act[i], cost_jinken_act[i], cost_hoken_act[i]])
        if total == 0:
            for arr in [cost_nenryo_act, cost_kotsu_act, cost_sharyo_act, cost_jinken_act, cost_hoken_act]:
                arr[i] = None

for i in range(12):
    items = [cost_nenryo_act[i], cost_kotsu_act[i], cost_sharyo_act[i],
             cost_jinken_act[i], cost_hoken_act[i]]
    if any(v is not None for v in items):
        ippan_cost_act.append(sum(v or 0 for v in items))
    else:
        ippan_cost_act.append(None)

ippan_gross_act = [
    (ippan_sales_act[i] - ippan_cost_act[i]) if ippan_sales_act[i] is not None and ippan_cost_act[i] is not None else None
    for i in range(12)
]
riyo_gross_act = [
    (riyo_sales_act[i] - riyo_cost_act[i]) if riyo_sales_act[i] is not None and riyo_cost_act[i] is not None else None
    for i in range(12)
]

# === 予算データ（index.html から転記） ===
ippan_sales_bud = [40260,41100,52700,55550,52000,58850,62850,56800,62850,62850,62850,56800]
ippan_cost_bud  = [31055,32419,40040,42303,39457,45086,48168,43105,48168,48168,48168,43105]
ippan_gross_bud = [ippan_sales_bud[i] - ippan_cost_bud[i] for i in range(12)]
riyo_sales_bud  = [16100]*12
riyo_cost_bud   = [15014]*12
riyo_gross_bud  = [riyo_sales_bud[i] - riyo_cost_bud[i] for i in range(12)]
cost_nenryo_bud = [9710,9830,13110,13890,12060,14470,15290,12840,15290,15290,15290,12840]
cost_kotsu_bud  = [4405,4405,5015,5295,5545,5595,6015,6015,6015,6015,6015,6015]
cost_sharyo_bud = [2250,2250,2270,2380,2550,2590,2780,2780,2780,2780,2780,2780]
cost_jinken_bud = [14082,15316,18722,19755,18282,21414,23003,20370,23003,23003,23003,20370]
cost_hoken_bud  = [608,618,923,983,1020,1017,1080,1100,1080,1080,1080,1100]
CF_FIXED = -16500

# === Excel 生成 ===
wb = openpyxl.Workbook()

# スタイル
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHDR_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONTH_LABELS = [f"{m}月" for m in months]


def fmt(v):
    return v if v is not None else "—"


def apply_header(ws, row, labels, fill=HEADER_FILL, font=HEADER_FONT):
    for i, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = fill
        c.font = font
        c.alignment = CENTER
        c.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --- Summary シート ---
ws0 = wb.active
ws0.title = "Summary"
ws0["A1"] = "フロンタル 予算実績レポート"
ws0["A1"].font = Font(bold=True, size=16, color="1E3A5F")
ws0["A2"] = f"作成日: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws0["A3"] = "単位: 千円"
ws0["A4"] = f"データソース: DriveDoor（2026/01/01〜{datetime.now().strftime('%Y/%m/%d')}取得）"
ws0["A5"] = f"CF_FIXED = {CF_FIXED:,}（固定費6,000 + リース7,500 + 金融3,000）"

ws0["A7"] = "■ 累計 KPI（1〜3月確定）"
ws0["A7"].font = Font(bold=True, size=12, color="1E3A5F")

kpi_headers = ["項目", "予算", "実績", "達成率", "差異"]
apply_header(ws0, 8, kpi_headers)

def safe_sum(arr, indices):
    return sum(arr[i] or 0 for i in indices if arr[i] is not None)

confirmed_idx = [0, 1, 2]
rows_kpi = [
    ("一般売上", safe_sum(ippan_sales_bud, confirmed_idx), safe_sum(ippan_sales_act, confirmed_idx)),
    ("一般原価", safe_sum(ippan_cost_bud, confirmed_idx), safe_sum(ippan_cost_act, confirmed_idx)),
    ("一般粗利", safe_sum(ippan_gross_bud, confirmed_idx), safe_sum(ippan_gross_act, confirmed_idx)),
    ("利用売上", safe_sum(riyo_sales_bud, confirmed_idx), safe_sum(riyo_sales_act, confirmed_idx)),
    ("外注費",   safe_sum(riyo_cost_bud, confirmed_idx),  safe_sum(riyo_cost_act, confirmed_idx)),
    ("利用粗利", safe_sum(riyo_gross_bud, confirmed_idx), safe_sum(riyo_gross_act, confirmed_idx)),
]
for i, (lbl, bud, act) in enumerate(rows_kpi, 9):
    ws0.cell(row=i, column=1, value=lbl).alignment = CENTER
    ws0.cell(row=i, column=2, value=bud).alignment = RIGHT
    ws0.cell(row=i, column=3, value=act).alignment = RIGHT
    ws0.cell(row=i, column=4, value=f"{act/bud*100:.1f}%" if bud else "—").alignment = RIGHT
    ws0.cell(row=i, column=5, value=act - bud).alignment = RIGHT
    for col in range(1, 6):
        ws0.cell(row=i, column=col).border = BORDER

set_col_widths(ws0, [15, 14, 14, 12, 14])


# --- 共通書き出し関数（予算→実績の順） ---
def write_buddet_table(ws, title, bud_arr, act_arr, gross_bud=None, gross_act=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    headers = ["月", "予算", "実績", "差異", "達成率"]
    if gross_bud is not None:
        headers += ["粗利 予算", "粗利 実績", "粗利率"]
    apply_header(ws, 3, headers)
    for i, m in enumerate(MONTH_LABELS):
        row = 4 + i
        bud = bud_arr[i]
        act = act_arr[i]
        diff = (act - bud) if act is not None else None
        rate = f"{act/bud*100:.1f}%" if act is not None and bud else "—"
        ws.cell(row=row, column=1, value=m).alignment = CENTER
        ws.cell(row=row, column=2, value=bud).alignment = RIGHT
        ws.cell(row=row, column=3, value=fmt(act)).alignment = RIGHT
        ws.cell(row=row, column=4, value=fmt(diff)).alignment = RIGHT
        ws.cell(row=row, column=5, value=rate).alignment = RIGHT
        if gross_bud is not None:
            gb = gross_bud[i]
            ga = gross_act[i]
            gr = f"{ga/act*100:.1f}%" if ga is not None and act else "—"
            ws.cell(row=row, column=6, value=gb).alignment = RIGHT
            ws.cell(row=row, column=7, value=fmt(ga)).alignment = RIGHT
            ws.cell(row=row, column=8, value=gr).alignment = RIGHT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = BORDER
    set_col_widths(ws, [8, 12, 12, 12, 10] + ([12, 12, 10] if gross_bud else []))


# --- ① 一般売上・粗利 ---
ws1 = wb.create_sheet("①一般売上・粗利")
write_buddet_table(ws1, "① 一般売上・粗利（予算 vs 実績）",
                   ippan_sales_bud, ippan_sales_act,
                   ippan_gross_bud, ippan_gross_act)

# --- ② 一般原価（費目別） ---
ws2 = wb.create_sheet("②一般原価費目別")
ws2["A1"] = "② 一般原価（費目別）予算 vs 実績"
ws2["A1"].font = Font(bold=True, size=14, color="1E3A5F")
cat_labels = ["燃料費", "交通費", "車両費", "保険料", "労務費", "合計"]
cat_act = [cost_nenryo_act, cost_kotsu_act, cost_sharyo_act, cost_hoken_act, cost_jinken_act]
cat_bud = [cost_nenryo_bud, cost_kotsu_bud, cost_sharyo_bud, cost_hoken_bud, cost_jinken_bud]

# ヘッダー2段
ws2.cell(row=3, column=1, value="月").fill = HEADER_FILL
ws2.cell(row=3, column=1).font = HEADER_FONT
ws2.cell(row=3, column=1).alignment = CENTER
ws2.cell(row=3, column=1).border = BORDER
ws2.merge_cells(start_row=3, end_row=4, start_column=1, end_column=1)
for i, lbl in enumerate(cat_labels):
    col_start = 2 + i * 3
    ws2.cell(row=3, column=col_start, value=lbl).fill = HEADER_FILL
    ws2.cell(row=3, column=col_start).font = HEADER_FONT
    ws2.cell(row=3, column=col_start).alignment = CENTER
    ws2.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start + 2)
    for j, sub in enumerate(["予算", "実績", "率"]):
        c = ws2.cell(row=4, column=col_start + j, value=sub)
        c.fill = SUBHDR_FILL
        c.font = Font(bold=True, size=10)
        c.alignment = CENTER
        c.border = BORDER
    for j in range(3):
        ws2.cell(row=3, column=col_start + j).border = BORDER

for i, m in enumerate(MONTH_LABELS):
    row = 5 + i
    ws2.cell(row=row, column=1, value=m).alignment = CENTER
    ws2.cell(row=row, column=1).border = BORDER
    tot_act = 0
    tot_bud = 0
    tot_act_valid = True
    for k, (a, b) in enumerate(zip(cat_act, cat_bud)):
        col_start = 2 + k * 3
        av = a[i]
        bv = b[i]
        tot_bud += bv
        if av is None:
            tot_act_valid = False
        else:
            tot_act += av
        rate = f"{av/bv*100:.1f}%" if av is not None and bv else "—"
        ws2.cell(row=row, column=col_start, value=bv).alignment = RIGHT
        ws2.cell(row=row, column=col_start + 1, value=fmt(av)).alignment = RIGHT
        ws2.cell(row=row, column=col_start + 2, value=rate).alignment = RIGHT
        for j in range(3):
            ws2.cell(row=row, column=col_start + j).border = BORDER
    # 合計列
    col_start = 2 + 5 * 3
    ws2.cell(row=row, column=col_start, value=tot_bud).alignment = RIGHT
    ws2.cell(row=row, column=col_start + 1, value=tot_act if tot_act_valid else "—").alignment = RIGHT
    ws2.cell(row=row, column=col_start + 2,
             value=f"{tot_act/tot_bud*100:.1f}%" if tot_act_valid and tot_bud else "—").alignment = RIGHT
    for j in range(3):
        ws2.cell(row=row, column=col_start + j).border = BORDER

set_col_widths(ws2, [8] + [10] * 18)


# --- ③ 利用売上・粗利 ---
ws3 = wb.create_sheet("③利用売上・粗利")
ws3["A1"] = "③ 利用売上・粗利（予算 vs 実績）"
ws3["A1"].font = Font(bold=True, size=14, color="1E3A5F")
headers3 = ["月", "売上 予算", "売上 実績", "差異", "達成率", "外注費 実績", "粗利 予算", "粗利 実績"]
apply_header(ws3, 3, headers3)
for i, m in enumerate(MONTH_LABELS):
    row = 4 + i
    sb = riyo_sales_bud[i]
    sa = riyo_sales_act[i]
    ca = riyo_cost_act[i]
    gb = riyo_gross_bud[i]
    ga = riyo_gross_act[i]
    diff = (sa - sb) if sa is not None else None
    rate = f"{sa/sb*100:.1f}%" if sa is not None and sb else "—"
    ws3.cell(row=row, column=1, value=m).alignment = CENTER
    ws3.cell(row=row, column=2, value=sb).alignment = RIGHT
    ws3.cell(row=row, column=3, value=fmt(sa)).alignment = RIGHT
    ws3.cell(row=row, column=4, value=fmt(diff)).alignment = RIGHT
    ws3.cell(row=row, column=5, value=rate).alignment = RIGHT
    ws3.cell(row=row, column=6, value=fmt(ca)).alignment = RIGHT
    ws3.cell(row=row, column=7, value=gb).alignment = RIGHT
    ws3.cell(row=row, column=8, value=fmt(ga)).alignment = RIGHT
    for col in range(1, 9):
        ws3.cell(row=row, column=col).border = BORDER
set_col_widths(ws3, [8, 12, 12, 12, 10, 12, 12, 12])


# --- ④ キャッシュフロー ---
ws4 = wb.create_sheet("④キャッシュフロー")
ws4["A1"] = "④ キャッシュフロー（概算）"
ws4["A1"].font = Font(bold=True, size=14, color="1E3A5F")
ws4["A2"] = f"CF_FIXED = {CF_FIXED:,} 千円/月（固定費6,000 + リース7,500 + 金融3,000）"
headers4 = ["月", "一般粗利", "利用粗利", "粗利合計", "固定費", "CF 予算", "CF 実績"]
apply_header(ws4, 4, headers4)
for i, m in enumerate(MONTH_LABELS):
    row = 5 + i
    ig = ippan_gross_act[i]
    rg = riyo_gross_act[i]
    g = (ig + rg) if ig is not None and rg is not None else None
    cf = (g + CF_FIXED) if g is not None else None
    cfb = ippan_gross_bud[i] + riyo_gross_bud[i] + CF_FIXED
    ws4.cell(row=row, column=1, value=m).alignment = CENTER
    ws4.cell(row=row, column=2, value=fmt(ig)).alignment = RIGHT
    ws4.cell(row=row, column=3, value=fmt(rg)).alignment = RIGHT
    ws4.cell(row=row, column=4, value=fmt(g)).alignment = RIGHT
    ws4.cell(row=row, column=5, value=CF_FIXED).alignment = RIGHT
    ws4.cell(row=row, column=6, value=cfb).alignment = RIGHT
    ws4.cell(row=row, column=7, value=fmt(cf)).alignment = RIGHT
    for col in range(1, 8):
        ws4.cell(row=row, column=col).border = BORDER
set_col_widths(ws4, [8, 12, 12, 12, 12, 12, 12])


# --- ⑤ 累計進捗 ---
ws5 = wb.create_sheet("⑤累計進捗")
ws5["A1"] = "⑤ 累計進捗（1〜3月確定）"
ws5["A1"].font = Font(bold=True, size=14, color="1E3A5F")
apply_header(ws5, 3, ["項目", "予算", "実績", "達成率", "差異"])
rows5 = [
    ("一般売上", ippan_sales_bud, ippan_sales_act),
    ("一般原価", ippan_cost_bud,  ippan_cost_act),
    ("一般粗利", ippan_gross_bud, ippan_gross_act),
    ("利用売上", riyo_sales_bud,  riyo_sales_act),
    ("外注費",   riyo_cost_bud,   riyo_cost_act),
    ("利用粗利", riyo_gross_bud,  riyo_gross_act),
]
for i, (lbl, bud, act) in enumerate(rows5):
    row = 4 + i
    b = safe_sum(bud, confirmed_idx)
    a = safe_sum(act, confirmed_idx)
    ws5.cell(row=row, column=1, value=lbl).alignment = CENTER
    ws5.cell(row=row, column=2, value=b).alignment = RIGHT
    ws5.cell(row=row, column=3, value=a).alignment = RIGHT
    ws5.cell(row=row, column=4, value=f"{a/b*100:.1f}%" if b else "—").alignment = RIGHT
    ws5.cell(row=row, column=5, value=a - b).alignment = RIGHT
    for col in range(1, 6):
        ws5.cell(row=row, column=col).border = BORDER
set_col_widths(ws5, [15, 14, 14, 12, 14])


# ======================================================
# 拠点別・車両別 集計用に日報Excelを再読込
# ======================================================
import openpyxl as _oxl
from datetime import datetime as _dt


def read_nippo_detail(filepath, office_type):
    """日報Excelから明細単位で読み取り、拠点別・車両別集計用のレコードを返す。
    各レコード = dict(office, vehicle, haisha, yosha_name, month, sales, cost)
    office_type: 'honsha' / 'kyoto' / 'fjs'
    """
    wb = _oxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[0]
        if date_val is None:
            continue
        if isinstance(date_val, _dt):
            month = date_val.month
        elif isinstance(date_val, str):
            try:
                month = _dt.strptime(date_val.strip(), "%Y/%m/%d").month
            except ValueError:
                continue
        else:
            continue
        haisha = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        vehicle = str(row[27]).strip() if len(row) > 27 and row[27] else '（未指定）'
        yosha_name = str(row[40]).strip() if len(row) > 40 and row[40] else ''
        sales = float(row[89]) if len(row) > 89 and row[89] is not None else 0
        cost = float(row[100]) if len(row) > 100 and row[100] is not None else 0

        # 拠点判定（FJS振替ルール）
        if office_type == 'fjs':
            if haisha == '自社':
                office = '京都（FJS自社→振替）'
                category = 'ippan'
            elif haisha == '傭車' and 'シクロ' in yosha_name:
                office = '京都（FJSシクロ→振替）'
                category = 'riyo'
            else:
                continue  # 除外
        elif office_type == 'kyoto':
            if haisha == '自社':
                office = '京都'
                category = 'ippan'
            else:
                continue  # 京都傭車は保留除外
        else:
            office = '本社'
            category = 'ippan' if haisha == '自社' else 'riyo'

        records.append({
            'office': office,
            'office_base': '本社' if office_type == 'honsha' else '京都',
            'vehicle': vehicle,
            'haisha': haisha,
            'category': category,
            'month': month,
            'sales': sales,
            'cost': cost,  # 傭車の場合のみ意味を持つ（自社はCSV由来）
        })
    wb.close()
    return records


def read_sharyo_detail(filepath):
    """車両経費CSVから車両別・拠点別の経費合計を返す。
    Returns: dict[(office_base, vehicle_name)] = total_cost (円)
    """
    import csv as _csv
    # 固定費除外（aggregate.pyと同じ）
    INCLUDE_CATS = {'燃料', '通行料', '固定車両費', 'タイヤ', 'オイル', '修繕費', '保険料', '法定福利費', '諸経費'}
    result = {}
    with open(filepath, encoding='utf-8-sig') as f:
        reader = _csv.DictReader(f)
        for row in reader:
            himoku = (row.get('車両整備経費区分') or '').strip()
            if himoku not in INCLUDE_CATS:
                continue
            office = (row.get('経費営業所') or '').strip()
            vehicle = (row.get('車両表示名') or '').strip()
            if '京都' in office:
                ob = '京都'
            else:
                ob = '本社'
            amount_str = (row.get('経費金額') or '0').strip().replace(',', '')
            try:
                amount = float(amount_str)
            except ValueError:
                amount = 0
            key = (ob, vehicle)
            result[key] = result.get(key, 0) + amount
    return result


# データ読み込み（明細）
honsha_records = read_nippo_detail(DATA_DIR / "nippo_honsha.xlsx", "honsha")
kyoto_records = read_nippo_detail(DATA_DIR / "nippo_kyoto.xlsx", "kyoto")
fjs_records = read_nippo_detail(DATA_DIR / "nippo_fjs.xlsx", "fjs")
all_records = honsha_records + kyoto_records + fjs_records
sharyo_vehicle = read_sharyo_detail(DATA_DIR / "sharyokeihi.csv")


# ======================================================
# ⑥ 拠点別粗利
# ======================================================
ws6 = wb.create_sheet("⑥拠点別粗利")
ws6["A1"] = "⑥ 拠点別 売上・原価・粗利（1〜3月確定、単位: 千円）"
ws6["A1"].font = Font(bold=True, size=14, color="1E3A5F")
ws6["A2"] = "※FJS自社は京都の一般に振替、FJSシクロシュプリームは京都の利用に振替。京都傭車は保留除外。"
ws6["A2"].font = Font(size=9, color="666666")

headers6 = ["拠点", "区分", "売上", "原価", "粗利", "粗利率"]
apply_header(ws6, 4, headers6)

# 集計
office_agg = {}  # key=(office_base, category) value=dict(sales, cost)
for r in all_records:
    if r['month'] > 3:  # 1-3月確定分のみ
        continue
    key = (r['office_base'], r['category'])
    d = office_agg.setdefault(key, {'sales': 0, 'cost': 0})
    d['sales'] += r['sales']
    # 利用貨物は支払金額が原価、一般貨物は後で車両経費から算出
    if r['category'] == 'riyo':
        d['cost'] += r['cost']

# 一般貨物の原価（車両経費CSVから拠点別合計、1-3月分）
# 注: sharyo_vehicle は累計なので、月次で分けるには read_sharyo_keihi を使うべき
# ここでは簡易的に 1-3月分が全量として扱う（4月以降のデータが入ってくるまでは問題ない想定）
# より正確には月別集計が必要。今回は sharyo_vehicle の合計を拠点別に分配
honsha_cost_ippan = sum(v for (ob, _), v in sharyo_vehicle.items() if ob == '本社')
kyoto_cost_ippan = sum(v for (ob, _), v in sharyo_vehicle.items() if ob == '京都')
# 一般原価を office_agg に反映
office_agg.setdefault(('本社', 'ippan'), {'sales': 0, 'cost': 0})['cost'] += honsha_cost_ippan
office_agg.setdefault(('京都', 'ippan'), {'sales': 0, 'cost': 0})['cost'] += kyoto_cost_ippan

row_i = 5
CATEGORY_LABEL = {'ippan': '一般貨物', 'riyo': '利用貨物'}
totals_by_office = {}
for ob in ['本社', '京都']:
    office_subtotal = {'sales': 0, 'cost': 0}
    for cat in ['ippan', 'riyo']:
        key = (ob, cat)
        if key not in office_agg:
            continue
        d = office_agg[key]
        sales_k = round(d['sales'] / 1000)
        cost_k = round(d['cost'] / 1000)
        gross = sales_k - cost_k
        rate = f"{gross/sales_k*100:.1f}%" if sales_k else "—"
        ws6.cell(row=row_i, column=1, value=ob).alignment = CENTER
        ws6.cell(row=row_i, column=2, value=CATEGORY_LABEL[cat]).alignment = CENTER
        ws6.cell(row=row_i, column=3, value=sales_k).alignment = RIGHT
        ws6.cell(row=row_i, column=4, value=cost_k).alignment = RIGHT
        ws6.cell(row=row_i, column=5, value=gross).alignment = RIGHT
        ws6.cell(row=row_i, column=6, value=rate).alignment = RIGHT
        for col in range(1, 7):
            ws6.cell(row=row_i, column=col).border = BORDER
        office_subtotal['sales'] += sales_k
        office_subtotal['cost'] += cost_k
        row_i += 1
    # 拠点小計
    gs = office_subtotal['sales']
    gc = office_subtotal['cost']
    gg = gs - gc
    ws6.cell(row=row_i, column=1, value=ob).alignment = CENTER
    ws6.cell(row=row_i, column=2, value='小計').alignment = CENTER
    ws6.cell(row=row_i, column=3, value=gs).alignment = RIGHT
    ws6.cell(row=row_i, column=4, value=gc).alignment = RIGHT
    ws6.cell(row=row_i, column=5, value=gg).alignment = RIGHT
    ws6.cell(row=row_i, column=6, value=f"{gg/gs*100:.1f}%" if gs else "—").alignment = RIGHT
    for col in range(1, 7):
        c = ws6.cell(row=row_i, column=col)
        c.border = BORDER
        c.fill = SUBHDR_FILL
        c.font = Font(bold=True)
    totals_by_office[ob] = office_subtotal
    row_i += 1

set_col_widths(ws6, [22, 12, 14, 14, 14, 10])


# ======================================================
# ⑦ 車両別粗利（自社車両のみ、拠点ごと降順）
# ======================================================
ws7 = wb.create_sheet("⑦車両別粗利")
ws7["A1"] = "⑦ 車両別 売上・原価・粗利（1〜3月確定、自社車両のみ、単位: 千円）"
ws7["A1"].font = Font(bold=True, size=14, color="1E3A5F")
ws7["A2"] = "※売上=請求金額合計（自社便のみ）／原価=車両経費CSVより該当車両合計／粗利=売上−原価。拠点内で粗利降順。"
ws7["A2"].font = Font(size=9, color="666666")

headers7 = ["拠点", "自動車表示名", "売上", "原価", "粗利", "粗利率"]
apply_header(ws7, 4, headers7)

# 車両別集計（自社のみ）
vehicle_sales = {}  # (office_base, vehicle) -> sales(円)
for r in all_records:
    if r['category'] != 'ippan':
        continue
    if r['month'] > 3:
        continue
    key = (r['office_base'], r['vehicle'])
    vehicle_sales[key] = vehicle_sales.get(key, 0) + r['sales']

# 原価は sharyo_vehicle から。全車両のユニーク集合を作る
all_vehicles = set(vehicle_sales.keys()) | set(sharyo_vehicle.keys())

vehicle_rows = []
for key in all_vehicles:
    ob, veh = key
    sales = vehicle_sales.get(key, 0)
    cost = sharyo_vehicle.get(key, 0)
    sales_k = round(sales / 1000)
    cost_k = round(cost / 1000)
    gross = sales_k - cost_k
    vehicle_rows.append({
        'office': ob,
        'vehicle': veh,
        'sales': sales_k,
        'cost': cost_k,
        'gross': gross,
    })

# 拠点ごとに粗利降順
row_i = 5
for ob in ['本社', '京都']:
    rows_in_office = [v for v in vehicle_rows if v['office'] == ob]
    rows_in_office.sort(key=lambda x: x['gross'], reverse=True)
    if not rows_in_office:
        continue
    # 拠点ヘッダ
    cell = ws7.cell(row=row_i, column=1, value=f"== {ob} ==")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    ws7.merge_cells(start_row=row_i, end_row=row_i, start_column=1, end_column=6)
    row_i += 1
    # 明細
    total_sales = total_cost = 0
    for v in rows_in_office:
        ws7.cell(row=row_i, column=1, value=v['office']).alignment = CENTER
        ws7.cell(row=row_i, column=2, value=v['vehicle']).alignment = Alignment(horizontal="left", vertical="center")
        ws7.cell(row=row_i, column=3, value=v['sales']).alignment = RIGHT
        ws7.cell(row=row_i, column=4, value=v['cost']).alignment = RIGHT
        ws7.cell(row=row_i, column=5, value=v['gross']).alignment = RIGHT
        rate = f"{v['gross']/v['sales']*100:.1f}%" if v['sales'] else "—"
        ws7.cell(row=row_i, column=6, value=rate).alignment = RIGHT
        for col in range(1, 7):
            ws7.cell(row=row_i, column=col).border = BORDER
        total_sales += v['sales']
        total_cost += v['cost']
        row_i += 1
    # 拠点合計
    total_gross = total_sales - total_cost
    ws7.cell(row=row_i, column=1, value=ob).alignment = CENTER
    ws7.cell(row=row_i, column=2, value='合計').alignment = CENTER
    ws7.cell(row=row_i, column=3, value=total_sales).alignment = RIGHT
    ws7.cell(row=row_i, column=4, value=total_cost).alignment = RIGHT
    ws7.cell(row=row_i, column=5, value=total_gross).alignment = RIGHT
    ws7.cell(row=row_i, column=6,
             value=f"{total_gross/total_sales*100:.1f}%" if total_sales else "—").alignment = RIGHT
    for col in range(1, 7):
        c = ws7.cell(row=row_i, column=col)
        c.border = BORDER
        c.fill = SUBHDR_FILL
        c.font = Font(bold=True)
    row_i += 2  # 空行

set_col_widths(ws7, [10, 28, 14, 14, 14, 10])


# --- 保存 ---
# 出力先: Dropbox の outbox フォルダ
OUTBOX = Path.home() / "Dropbox" / "kuroda_work" / "outbox"
OUTBOX.mkdir(parents=True, exist_ok=True)
out_path = OUTBOX / f"frontal_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(out_path)
print(f"Excel report saved: {out_path}")

# バックアップとして reports/ にも保存
backup_path = BASE / "reports" / f"frontal_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
backup_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(backup_path)
print(f"Backup saved: {backup_path}")
